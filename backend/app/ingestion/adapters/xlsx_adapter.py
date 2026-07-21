"""XLSX document adapter.

Extracts structured data from Excel files. Each sheet
is treated as a separate page with its own table.
"""
from __future__ import annotations

import logging
from pathlib import Path

from app.ingestion.adapters.base import (
    DocumentAdapter,
    ExtractedPage,
    ExtractionResult,
)

logger = logging.getLogger(__name__)


class XLSXAdapter(DocumentAdapter):
    """Extracts content from Excel (XLSX) files."""

    @property
    def supported_extensions(self) -> tuple[str, ...]:
        return (".xlsx", ".xls")

    def can_handle(self, file_path: Path) -> bool:
        return file_path.suffix.lower() in self.supported_extensions

    def extract(self, file_path: Path) -> ExtractionResult:
        try:
            return self._extract_with_openpyxl(file_path)
        except ImportError:
            logger.info("openpyxl not available")
            return ExtractionResult(
                pages=(),
                total_pages=0,
                extraction_method="failed",
                errors=("openpyxl not installed",),
            )
        except Exception as e:
            return ExtractionResult(
                pages=(),
                total_pages=0,
                extraction_method="failed",
                errors=(str(e),),
            )

    def _extract_with_openpyxl(self, file_path: Path) -> ExtractionResult:
        from openpyxl import load_workbook

        wb = load_workbook(str(file_path), read_only=True, data_only=True)
        pages: list[ExtractedPage] = []

        for sheet_idx, sheet_name in enumerate(wb.sheetnames):
            ws = wb[sheet_name]
            rows: list[list[str]] = []
            for row in ws.iter_rows(values_only=True):
                string_row = [str(cell) if cell is not None else "" for cell in row]
                if any(cell.strip() for cell in string_row):
                    rows.append(string_row)

            if not rows:
                continue

            # Build text representation
            header = rows[0] if rows else []
            text_lines = [f"Sheet: {sheet_name}", ", ".join(header)]
            for row in rows[1:]:
                record = ", ".join(
                    f"{h}: {v}" for h, v in zip(header, row) if v.strip()
                )
                if record:
                    text_lines.append(record)

            pages.append(ExtractedPage(
                page_number=sheet_idx + 1,
                text="\n".join(text_lines),
                tables=(rows,),
                metadata={
                    "sheet_name": sheet_name,
                    "row_count": len(rows) - 1,
                    "columns": header,
                },
            ))

        wb.close()

        return ExtractionResult(
            pages=tuple(pages),
            total_pages=len(pages),
            extraction_method="openpyxl",
        )
